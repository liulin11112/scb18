'''
接口自动化测试
1、excel测试用例准备好，代码可以自动读取用例数据
2、调用接口函数，执行用例，得到响应结果
3、断言：响应结果与预期结果判断         ---通过/不通过
4、Excel中回写测试结果
'''

import openpyxl
import  requests
import jsonpath

#读取Excel中测试用例的数据，封装成一个函数
def read_case(filename,sheetname):
    wb = openpyxl.load_workbook(filename)   #加载工作簿，打开一个Excel文件
    sheet = wb[sheetname]      #打开某一个表单
    row_max = sheet.max_row   #获取最大行数

    case_list = []      #新建空列表，存放for循环依次读取到的测试用例数据
    for i in range(2,row_max+1):
        # cell = sheet.cell(row = 1,column = 1).value     #获取单元格的值
        data_dict = dict(case_id = sheet.cell(row = i,column = 1).value,      #读取case_id值
        url = sheet.cell(row = i,column = 5).value,      #读取url值
        data = sheet.cell(row = i,column = 6).value,     #读取data
        expect = sheet.cell(row = i,column = 7).value)   #读取期望值
        case_list.append(data_dict)     #把每一行读取到的测试用例数据生成的字典，追加到list中
    return case_list

#调用函数
# case_data = read_case('test_case_api.xlsx','login')
# print(case_data)


#写入结果到excel，定义成一个函数
def write_result(filename,sheetname,row,column,final_result):
    wb = openpyxl.load_workbook(filename)    #加载工作簿，打开一个Excel文件
    sheet = wb[sheetname]  # 打开某一个表单
    sheet.cell(row=row,column=column).value=final_result
    wb.save(filename)

# write_result('../test_data/test_case_api.xlsx', 'login', 2, 8, 'true')
# result = openpyxl.load_workbook('../test_data/test_case_api.xlsx')['login'].cell(row = 2, column = 8).value
# print(result)


#把接口请求封装成函数
def api_fun(url,data):
    headers = {'X-Lemonban-Media-Type': 'lemonban.v2', 'Content-Type': 'application/json'}  # 请求头
    res_login = requests.post(url=url, json=data, headers=headers).json()
    return res_login

#调用api_fun()
url_login = 'http://120.78.128.25:8766/futureloan/member/login'     #请求头
data_login = {"mobile_phone": "13321860729","pwd": "lemon123456"}    #请求体
result = api_fun(url = url_login,data = data_login)
# print(result)
